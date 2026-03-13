#!/usr/bin/env python3
"""
backtest.py — Rebuild the model at historical cutoffs and score predictions.

Produces a JSON report with headline and district-level calibration metrics.
"""

import argparse
import json
import math
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = REPO_ROOT / "scripts"
SNAPSHOTS_DIR = REPO_ROOT / "data" / "snapshots"
OUTPUT_PATH = REPO_ROOT / "data" / "calibration.json"

THRESHOLDS = {
    1: 5238, 2: 4687, 3: 4737, 4: 5099, 5: 4115, 6: 4745, 7: 5294,
    8: 4910, 9: 4805, 10: 2975, 11: 4890, 12: 3248, 13: 4088, 14: 5680,
    15: 4596, 16: 4347, 17: 5368, 18: 5093, 19: 5715, 20: 5292, 21: 5684,
    22: 5411, 23: 4253, 24: 3857, 25: 4929, 26: 5178, 27: 5696, 28: 5437,
    29: 5382,
}
STATEWIDE_TARGET = 140748


def parse_counts(path: Path) -> dict[int, int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    counts = {d: 0 for d in THRESHOLDS}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        try:
            district = int(row[3])
        except (TypeError, ValueError):
            continue
        if district in counts:
            counts[district] += 1
    wb.close()
    return counts


def symlink_subset(files: list[Path], dest_dir: Path) -> None:
    dest_dir.mkdir(parents=True, exist_ok=True)
    for src in files:
        target = dest_dir / src.name
        if target.exists() or target.is_symlink():
            target.unlink()
        target.symlink_to(src)


def run_command(args: list[str]) -> None:
    subprocess.run(args, check=True, cwd=REPO_ROOT)


def brier(prob: float, actual: int) -> float:
    return (prob - actual) ** 2


def log_loss(prob: float, actual: int) -> float:
    clipped = min(max(prob, 1e-6), 1 - 1e-6)
    return -(actual * math.log(clipped) + (1 - actual) * math.log(1 - clipped))


def main() -> None:
    parser = argparse.ArgumentParser(description="Backtest the model across historical snapshots.")
    parser.add_argument("--snapshots-dir", type=Path, default=SNAPSHOTS_DIR)
    parser.add_argument("--output", type=Path, default=OUTPUT_PATH)
    parser.add_argument("--min-snapshots", type=int, default=3, help="Minimum number of snapshots before scoring a cutoff")
    args = parser.parse_args()

    files = sorted(args.snapshots_dir.glob("*.xlsx"), key=lambda p: p.stem)
    if len(files) < args.min_snapshots + 1:
        raise SystemExit("Not enough snapshots to run backtest.")

    final_counts = parse_counts(files[-1])
    final_total = sum(final_counts.values())
    final_district_rule_actual = int(sum(1 for d, count in final_counts.items() if count >= THRESHOLDS[d]) >= 26)
    final_statewide_actual = int(final_total >= STATEWIDE_TARGET)
    final_ballot_actual = int(final_district_rule_actual and final_statewide_actual)

    results = []

    with tempfile.TemporaryDirectory(prefix="deadreckoning-backtest-") as tmp_root:
        tmp_root_path = Path(tmp_root)
        for idx in range(args.min_snapshots - 1, len(files) - 1):
            subset = files[: idx + 1]
            cutoff = subset[-1]
            tmp_snapshots = tmp_root_path / "snapshots"
            tmp_history = tmp_root_path / "history.json"
            tmp_removals = tmp_root_path / "removals.json"
            tmp_output = tmp_root_path / "data.json"

            symlink_subset(subset, tmp_snapshots)

            run_command([
                sys.executable,
                str(SCRIPTS_DIR / "replay.py"),
                "--snapshots-dir", str(tmp_snapshots),
                "--history-out", str(tmp_history),
                "--removals-out", str(tmp_removals),
            ])
            run_command([
                sys.executable,
                str(SCRIPTS_DIR / "process.py"),
                "--file", str(cutoff),
                "--history", str(tmp_history),
                "--removals", str(tmp_removals),
                "--output", str(tmp_output),
            ])

            prediction = json.load(open(tmp_output))
            overall = prediction["overall"]
            district_rule_prob = overall.get("pDistrictRule", overall.get("pQualify", 0.0))
            ballot_prob = overall.get("pBallotQualified", district_rule_prob)
            statewide_prob = overall["statewideProjection"]["pReachTarget"]

            district_brier = 0.0
            for district in prediction["districts"]:
                actual = int(final_counts[district["d"]] >= THRESHOLDS[district["d"]])
                district_brier += brier(district["prob"], actual)
            district_brier /= len(prediction["districts"])

            results.append({
                "snapshot": cutoff.stem,
                "districtRuleProb": district_rule_prob,
                "ballotProb": ballot_prob,
                "statewideProb": statewide_prob,
                "districtRuleBrier": round(brier(district_rule_prob, final_district_rule_actual), 6),
                "districtRuleLogLoss": round(log_loss(district_rule_prob, final_district_rule_actual), 6),
                "ballotBrier": round(brier(ballot_prob, final_ballot_actual), 6),
                "ballotLogLoss": round(log_loss(ballot_prob, final_ballot_actual), 6),
                "statewideBrier": round(brier(statewide_prob, final_statewide_actual), 6),
                "statewideLogLoss": round(log_loss(statewide_prob, final_statewide_actual), 6),
                "districtMeanBrier": round(district_brier, 6),
            })

    summary = {
        "snapshotsScored": len(results),
        "finalOutcome": {
            "districtRuleQualified": bool(final_district_rule_actual),
            "statewideTargetReached": bool(final_statewide_actual),
            "ballotQualified": bool(final_ballot_actual),
        },
        "meanDistrictRuleBrier": round(sum(r["districtRuleBrier"] for r in results) / len(results), 6),
        "meanBallotBrier": round(sum(r["ballotBrier"] for r in results) / len(results), 6),
        "meanStatewideBrier": round(sum(r["statewideBrier"] for r in results) / len(results), 6),
        "meanDistrictMeanBrier": round(sum(r["districtMeanBrier"] for r in results) / len(results), 6),
    }

    report = {"summary": summary, "results": results}
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w") as f:
        json.dump(report, f, indent=2)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
