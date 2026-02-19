#!/usr/bin/env python3
"""
replay.py — Build data/history.json from all xlsx snapshots in data/snapshots/.

Each snapshot filename must be YYYY-MM-DD.xlsx (the date the LG published it).

Outputs data/history.json with:
  - per-snapshot verified counts per district
  - statewide totals
  - computed inter-snapshot deltas (new sigs, removals, net)
  - per-district rejection/removal rates

Usage:
    .venv/bin/python scripts/replay.py
"""

import json
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
SNAPSHOTS_DIR = REPO_ROOT / "data" / "snapshots"
HISTORY_PATH = REPO_ROOT / "data" / "history.json"

THRESHOLDS = {
    1: 5238, 2: 4687, 3: 4737, 4: 5099, 5: 4115, 6: 4745, 7: 5294,
    8: 4910, 9: 4805, 10: 2975, 11: 4890, 12: 3248, 13: 4088, 14: 5680,
    15: 4596, 16: 4347, 17: 5368, 18: 5093, 19: 5715, 20: 5292, 21: 5684,
    22: 5411, 23: 4253, 24: 3857, 25: 4929, 26: 5178, 27: 5696, 28: 5437,
    29: 5382,
}

CLERK_DEADLINE = date(2026, 3, 9)
SUBMISSION_DEADLINE = date(2026, 2, 15)  # last day petitioners could submit new sigs


def parse_xlsx_counts(path: Path) -> dict[int, int]:
    """Read xlsx, return {district: verified_count}. Fast — only reads col D."""
    print(f"  Parsing {path.name} ...", end=" ", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    counts: dict[int, int] = defaultdict(int)
    skipped = 0
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            skipped += 1
            continue
        district_raw = row[3]
        try:
            district = int(district_raw)
        except (TypeError, ValueError):
            skipped += 1
            continue
        if district not in THRESHOLDS:
            skipped += 1
            continue
        counts[district] += 1
        total += 1
    wb.close()
    # Fill in zeros for any missing districts
    for d in THRESHOLDS:
        if d not in counts:
            counts[d] = 0
    print(f"{total:,} rows")
    return dict(counts)


def days_to_deadline(snap_date: date) -> int:
    return max((CLERK_DEADLINE - snap_date).days, 0)


def linear_project(history_vals: list[tuple[date, int]], target_date: date) -> float:
    """
    Weighted linear regression on (days_since_first, count) pairs.
    Recent snapshots get exponentially higher weight so late-campaign surges
    are captured. Projects to target_date.
    """
    if len(history_vals) < 2:
        return float(history_vals[-1][1]) if history_vals else 0.0

    base_date = history_vals[0][0]
    xs = [(d - base_date).days for d, _ in history_vals]
    ys = [v for _, v in history_vals]
    n = len(xs)

    # Exponential weights: most recent point gets weight 1.0, oldest gets 0.3^(n-1)
    # This lets the model track recent acceleration without ignoring history entirely.
    decay = 0.75
    ws = [decay ** (n - 1 - i) for i in range(n)]
    W = sum(ws)

    # Weighted least squares
    xbar = sum(w * x for w, x in zip(ws, xs)) / W
    ybar = sum(w * y for w, y in zip(ws, ys)) / W
    num = sum(w * (x - xbar) * (y - ybar) for w, x, y in zip(ws, xs, ys))
    den = sum(w * (x - xbar) ** 2 for w, x in zip(ws, xs))

    if den == 0:
        return float(ys[-1])

    slope = num / den
    intercept = ybar - slope * xbar

    target_x = (target_date - base_date).days
    projected = intercept + slope * target_x

    # Never project below last known value (can only go up)
    return max(projected, float(ys[-1]))


def detect_anomalies(snapshots: list[dict], threshold_pct: float = 0.02) -> list[dict]:
    """
    Scan inter-snapshot deltas for anomalous drops that suggest packet-level
    fraud rejections rather than normal signature-by-signature corrections.

    A drop is flagged if a district loses more than `threshold_pct` of its
    previous count in a single interval (default: 2%).

    Returns a list of anomaly records, each with:
      date, district, prev_count, cur_count, drop, drop_pct
    """
    anomalies = []
    for i in range(1, len(snapshots)):
        prev = snapshots[i - 1]
        cur = snapshots[i]
        for d in THRESHOLDS:
            prev_count = prev["districts"].get(str(d), 0)
            cur_count = cur["districts"].get(str(d), 0)
            if prev_count == 0:
                continue
            drop = prev_count - cur_count
            drop_pct = drop / prev_count
            if drop > 0 and drop_pct >= threshold_pct:
                anomalies.append({
                    "date": cur["date"],
                    "district": d,
                    "prevCount": prev_count,
                    "curCount": cur_count,
                    "drop": drop,
                    "dropPct": round(drop_pct, 4),
                    "prevDate": prev["date"],
                })
    return sorted(anomalies, key=lambda x: x["dropPct"], reverse=True)


def compute_post_deadline_removal_rates(
    history_by_district: dict[int, list[tuple[date, int]]],
) -> dict[int, float]:
    """
    Compute removal rates using only post-submission-deadline snapshots.
    These are pure clerk-review removals with no new additions mixed in.

    Returns {district: removal_rate} where removal_rate = total_removals / peak.
    Returns 0.0 for districts with no post-deadline observations.
    """
    rates = {}
    for d, vals in history_by_district.items():
        post = [(dt, count) for dt, count in vals if dt > SUBMISSION_DEADLINE]
        if len(post) < 2:
            rates[d] = 0.0
            continue
        counts = [count for _, count in post]
        peak = max(counts)
        total_removals = sum(
            max(0, counts[i] - counts[i + 1])
            for i in range(len(counts) - 1)
        )
        rates[d] = round(total_removals / peak, 4) if peak > 0 else 0.0
    return rates


def compute_rejection_rate(history_by_district: dict[int, list[tuple[date, int]]]) -> dict[int, float]:
    """
    For each district, compute observed removal rate:
      removal_rate = total_removals / peak_verified
    where removals = sum of all inter-snapshot declines.
    """
    rates = {}
    for d, vals in history_by_district.items():
        if len(vals) < 2:
            rates[d] = 0.0
            continue
        counts = [v for _, v in vals]
        peak = max(counts)
        total_removals = sum(
            max(0, counts[i] - counts[i + 1])
            for i in range(len(counts) - 1)
        )
        rates[d] = round(total_removals / peak, 4) if peak > 0 else 0.0
    return rates


def main():
    if not SNAPSHOTS_DIR.exists():
        print(f"ERROR: {SNAPSHOTS_DIR} does not exist", file=sys.stderr)
        sys.exit(1)

    # Find all YYYY-MM-DD.xlsx files
    files = sorted(
        [f for f in SNAPSHOTS_DIR.glob("*.xlsx") if len(f.stem) == 10],
        key=lambda f: f.stem
    )

    if not files:
        print(f"No YYYY-MM-DD.xlsx files found in {SNAPSHOTS_DIR}", file=sys.stderr)
        sys.exit(1)

    print(f"Found {len(files)} snapshot(s):\n")

    snapshots = []
    history_by_district: dict[int, list[tuple[date, int]]] = defaultdict(list)

    for f in files:
        snap_date = date.fromisoformat(f.stem)
        counts = parse_xlsx_counts(f)
        statewide = sum(counts.values())
        snapshots.append({
            "date": f.stem,
            "total": statewide,
            "districts": {str(d): counts.get(d, 0) for d in sorted(THRESHOLDS.keys())},
        })
        for d in THRESHOLDS:
            history_by_district[d].append((snap_date, counts.get(d, 0)))

    print(f"\nBuilding inter-snapshot deltas...")

    # Compute inter-snapshot deltas for each snapshot (except the first)
    for i, snap in enumerate(snapshots):
        if i == 0:
            snap["deltas"] = {str(d): 0 for d in THRESHOLDS}
            snap["removals"] = {str(d): 0 for d in THRESHOLDS}
            snap["net"] = {str(d): 0 for d in THRESHOLDS}
            snap["totalDelta"] = 0
            snap["totalRemovals"] = 0
            continue
        prev = snapshots[i - 1]
        deltas = {}
        removals = {}
        net = {}
        for d in THRESHOLDS:
            cur = snap["districts"][str(d)]
            prv = prev["districts"][str(d)]
            diff = cur - prv
            # New signatures added (could be undercount if removals happened simultaneously)
            # Net is what we observe; we can't separate adds from removes cleanly
            # but we can flag net negatives as confirmed removals
            deltas[str(d)] = max(0, diff)    # observed net gains
            removals[str(d)] = max(0, -diff)  # observed net losses (removals/rejections)
            net[str(d)] = diff
        snap["deltas"] = deltas
        snap["removals"] = removals
        snap["net"] = net
        snap["totalDelta"] = snap["total"] - prev["total"]
        snap["totalRemovals"] = sum(removals.values())

    # Anomaly detection — packet-level fraud/rejection events
    print("Scanning for anomalous drops (≥2% single-interval decline)...")
    anomalies = detect_anomalies(snapshots, threshold_pct=0.02)
    if anomalies:
        print(f"  ⚠️  {len(anomalies)} anomalous drop(s) detected:")
        for a in anomalies:
            print(f"     D{a['district']} on {a['date']}: -{a['drop']:,} sigs ({a['dropPct']:.1%}) "
                  f"[{a['prevCount']:,} → {a['curCount']:,}]")
    else:
        print("  No anomalous drops found.")

    # Peak verified per district (highest count ever seen across all snapshots)
    peak_verified = {}
    for d in THRESHOLDS:
        vals = history_by_district[d]
        peak_verified[str(d)] = max((count for _, count in vals), default=0)

    # Post-deadline removal rates (pure clerk-review signal, no new additions)
    post_deadline_in_history = any(
        dt > SUBMISSION_DEADLINE
        for vals in history_by_district.values()
        for dt, _ in vals
    )
    if post_deadline_in_history:
        print("Computing post-submission-deadline removal rates...")
        post_deadline_rates = compute_post_deadline_removal_rates(history_by_district)
        statewide_post_deadline_rate = round(
            sum(post_deadline_rates.values()) / len(post_deadline_rates), 4
        )
    else:
        print("No post-deadline snapshots yet — post-deadline removal rates unavailable.")
        post_deadline_rates = {d: 0.0 for d in THRESHOLDS}
        statewide_post_deadline_rate = 0.0

    # Rejection rates per district (full history)
    print("Computing rejection/removal rates...")
    rejection_rates = compute_rejection_rate(history_by_district)

    # Statewide rejection rate
    statewide_rejection_rate = round(
        sum(rejection_rates.values()) / len(rejection_rates), 4
    ) if rejection_rates else 0.0

    # Linear projections to March 9 for each district
    print("Computing trajectory projections to March 9...")
    projections = {}
    for d in sorted(THRESHOLDS.keys()):
        vals = history_by_district[d]
        proj = linear_project(vals, CLERK_DEADLINE)
        # Apply rejection rate haircut to projection
        rejection_adjusted = proj * (1 - rejection_rates.get(d, 0))
        projections[str(d)] = {
            "raw": round(proj, 1),
            "rejectionAdjusted": round(rejection_adjusted, 1),
            "threshold": THRESHOLDS[d],
            "pctOfThreshold": round(rejection_adjusted / THRESHOLDS[d], 4) if THRESHOLDS[d] > 0 else 0,
        }

    # Statewide projection
    statewide_proj_raw = sum(v["raw"] for v in projections.values())
    statewide_proj_adj = sum(v["rejectionAdjusted"] for v in projections.values())

    # Velocity: signatures per day (last interval)
    last_snap = snapshots[-1]
    second_last = snapshots[-2] if len(snapshots) >= 2 else None
    if second_last:
        days_between = (
            date.fromisoformat(last_snap["date"]) - date.fromisoformat(second_last["date"])
        ).days or 1
        daily_velocity = last_snap["totalDelta"] / days_between
    else:
        daily_velocity = 0.0

    days_left = days_to_deadline(date.fromisoformat(last_snap["date"]))

    output = {
        "generated": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "snapshotCount": len(snapshots),
        "firstSnapshot": snapshots[0]["date"],
        "lastSnapshot": last_snap["date"],
        "daysToDeadline": days_left,
        "dailyVelocity": round(daily_velocity, 1),
        "statewideRejectionRate": statewide_rejection_rate,
        "projections": {
            "statewideRaw": round(statewide_proj_raw, 0),
            "statewideAdjusted": round(statewide_proj_adj, 0),
            "byDistrict": projections,
        },
        "rejectionRates": {str(d): rejection_rates[d] for d in sorted(THRESHOLDS.keys())},
        "postDeadlineRemovalRates": {str(d): post_deadline_rates[d] for d in sorted(THRESHOLDS.keys())},
        "statewidePostDeadlineRate": statewide_post_deadline_rate,
        "peakVerified": peak_verified,
        "postDeadlineDataAvailable": post_deadline_in_history,
        "anomalies": anomalies,
        "snapshots": snapshots,
    }

    HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(HISTORY_PATH, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\nWrote {HISTORY_PATH}")
    print(f"\n--- Summary ---")
    print(f"Snapshots: {snapshots[0]['date']} → {last_snap['date']}")
    print(f"Statewide verified (latest): {last_snap['total']:,}")
    print(f"Days to clerk deadline: {days_left}")
    print(f"Daily velocity (last interval): {daily_velocity:,.0f} sigs/day")
    print(f"Statewide rejection rate (observed): {statewide_rejection_rate:.1%}")
    print(f"Projected statewide (raw linear): {statewide_proj_raw:,.0f}")
    print(f"Projected statewide (rejection-adjusted): {statewide_proj_adj:,.0f}")
    print(f"\nPer-district rejection rates (top 5 highest):")
    top_removals = sorted(rejection_rates.items(), key=lambda x: x[1], reverse=True)[:5]
    for d, r in top_removals:
        print(f"  D{d}: {r:.1%}")


if __name__ == "__main__":
    main()
