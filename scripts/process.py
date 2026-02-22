#!/usr/bin/env python3
"""
process.py — Parses petition xlsx, computes district stats, writes public/data.json.

If data/history.json exists (built by replay.py), uses it for:
  - Real inter-snapshot velocity
  - Per-district rejection/removal rates
  - Linear trajectory projections to March 9

Usage:
    python scripts/process.py [--file path/to/file.xlsx]
"""

import argparse
import hashlib
import json
import math
import os
import sys
from collections import defaultdict
from datetime import datetime, date, timezone
from pathlib import Path

from dateutil import parser as dateutil_parser
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
PUBLIC_DIR = REPO_ROOT / "public"
LATEST_PATH = DATA_DIR / "latest.xlsx"
DATA_JSON_PATH = PUBLIC_DIR / "data.json"
LOOKUP_INDEX_PATH = PUBLIC_DIR / "lookup.json"
HISTORY_PATH = DATA_DIR / "history.json"

THRESHOLDS = {
    1: 5238, 2: 4687, 3: 4737, 4: 5099, 5: 4115, 6: 4745, 7: 5294,
    8: 4910, 9: 4805, 10: 2975, 11: 4890, 12: 3248, 13: 4088, 14: 5680,
    15: 4596, 16: 4347, 17: 5368, 18: 5093, 19: 5715, 20: 5292, 21: 5684,
    22: 5411, 23: 4253, 24: 3857, 25: 4929, 26: 5178, 27: 5696, 28: 5437,
    29: 5382,
}

# Only used if history.json is absent (fallback for first-run / no snapshot data)
ESTIMATED_VALID_UNVERIFIED = 0
QUALIFICATION_THRESHOLD_STATEWIDE = 140748
DISTRICTS_REQUIRED = 26
TOTAL_DISTRICTS = 29
CLERK_DEADLINE = date(2026, 3, 9)
CLERK_DEADLINE_STR = "2026-03-09"
SUBMISSION_DEADLINE = date(2026, 2, 15)  # last day petitioners could submit new sigs
ELECTION_DATE = "2026-11-03"

# Number of historical weekly buckets for sparkline display
WEEKLY_BUCKETS = 10


# ---------------------------------------------------------------------------
# Name lookup helpers
# ---------------------------------------------------------------------------

def normalize_name(raw: str) -> str:
    """
    Normalize a name from the xlsx for hashing.
    Input format: "Lastname, Firstname Middlename" or "Lastname, Firstname"
    Output: "LASTNAME,FIRSTNAME" (uppercase, no middle name, no spaces)
    """
    raw = str(raw).strip().upper()
    if ',' in raw:
        parts = raw.split(',', 1)
        last = parts[0].strip()
        first_parts = parts[1].strip().split()
        first = first_parts[0] if first_parts else ''
    else:
        last = raw
        first = ''
    return f"{last},{first}"


def name_hash(normalized: str) -> str:
    """First 20 hex chars of SHA-256 (10 bytes = 80 bits)."""
    return hashlib.sha256(normalized.encode('utf-8')).hexdigest()[:20]


def bloom_hash_positions(key: str, m: int, k: int) -> list[int]:
    """
    Generate k bit positions for a key in a bloom filter of m bits.
    Uses double hashing: h_i(x) = (h1(x) + i*h2(x)) mod m
    where h1 = SHA-256 first 8 bytes, h2 = SHA-256 bytes 8-16.
    This matches the JS implementation exactly.
    """
    digest = hashlib.sha256(key.encode('utf-8')).digest()
    h1 = int.from_bytes(digest[0:8], 'big')
    h2 = int.from_bytes(digest[8:16], 'big') | 1  # ensure odd for full coverage
    return [(h1 + i * h2) % m for i in range(k)]


def build_bloom_filter(keys: list[str], m: int = 65536, k: int = 7) -> dict:
    """
    Build a bloom filter for a set of keys.
    m: number of bits (65536 = 8KB per district)
    k: number of hash functions (7 is optimal for ~3000-8000 entries at 8KB)
    Returns a base64-encoded bit array.
    """
    import base64
    bits = bytearray(m // 8)
    for key in keys:
        for pos in bloom_hash_positions(key, m, k):
            bits[pos >> 3] |= (1 << (pos & 7))
    return {
        "m": m,
        "k": k,
        "bits": base64.b64encode(bytes(bits)).decode('ascii'),
    }


def build_lookup_index(district_names: dict[int, list[str]]) -> dict:
    """
    Build per-district bloom filters. Key = "LASTNAME,FIRSTNAME,D{n}".
    This scopes lookups to district so cross-district false positives are eliminated.
    """
    import base64
    M = 65536  # 8KB per district (64K bits)
    K = 7      # 7 hash functions → ~0.3% FP rate at 8K entries

    districts_bloom = {}
    total_names = 0
    for d_num, names in district_names.items():
        keys = []
        for raw in names:
            norm = normalize_name(raw)
            if norm and norm != ',':
                # District-scoped key: eliminates cross-district false positives
                keys.append(f"{norm},D{d_num}")
        bf = build_bloom_filter(keys, M, K)
        districts_bloom[str(d_num)] = bf
        total_names += len(keys)

    return {
        "version": 2,
        "m": M,
        "k": K,
        "count": total_names,
        "districts": districts_bloom,
    }


# ---------------------------------------------------------------------------
# History loader
# ---------------------------------------------------------------------------

def load_history() -> dict | None:
    """Load data/history.json if present, else return None."""
    if not HISTORY_PATH.exists():
        return None
    try:
        with open(HISTORY_PATH) as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: could not load history.json: {e}")
        return None


# ---------------------------------------------------------------------------
# Probability model
# ---------------------------------------------------------------------------

def compute_district_prob(
    verified: int,
    threshold: int,
    trend: str,
    final_week_sigs: int,
    projected_adj: float,   # rejection-adjusted linear projection to deadline
    rejection_rate: float,  # observed per-district removal rate [0,1]
) -> float:
    """
    Compute P(district meets threshold) using history-informed inputs where available.
    Returns a value in [0.00, 1.00].
    """
    if verified >= threshold:
        return 1.0

    # Structural feasibility gate: if the projection can't get close to threshold,
    # the district has very low odds regardless of trend noise.
    proj_pct = projected_adj / threshold if threshold > 0 else 0.0
    if proj_pct < 0.65:
        # Very far off — scale from 0 to 0.02
        return round(proj_pct * 0.031, 4)  # max 0.02 at 0.65
    if proj_pct < 0.80:
        # Hard climb zone: 0.02 → 0.08
        return round(0.02 + (proj_pct - 0.65) * 0.40, 4)
    if proj_pct < 0.90:
        # Getting realistic: 0.08 → 0.18
        return round(0.08 + (proj_pct - 0.80) * 1.00, 4)

    # Base: raw fraction already verified
    base_score = verified / threshold

    # Trend multiplier
    trend_mult = {"ACCEL": 1.08, "STABLE": 1.0, "DECEL": 0.90}.get(trend, 1.0)

    # Rejection penalty: high removal rate districts get a downward nudge
    rejection_penalty = rejection_rate * 0.5  # 10% removal rate → -0.05 penalty

    # Weighted combination
    raw = 0.45 * base_score + 0.35 * proj_pct + 0.20 * (base_score * trend_mult)
    raw -= rejection_penalty

    # Sigmoid-style squeeze
    if base_score < 0.50:
        raw *= 0.60
    elif base_score < 0.75:
        raw *= 0.85
    elif base_score >= 0.95:
        raw = max(raw, 0.85)

    # Velocity bonus
    if final_week_sigs > 500:
        raw += 0.03
    elif final_week_sigs > 200:
        raw += 0.01

    return max(0.00, min(0.99, raw))


def bayesian_removal_rate(
    observed_rate: float,
    days_post_deadline: int,
    clerk_window_days: int = 22,  # Feb 15 → Mar 9
) -> float:
    """
    Blend observed post-deadline removal rate with an empirical prior.

    Prior: 1.65% total removal over the full clerk window, derived from
    county clerk removal-request data reported by KSL/SLTrib (Feb 12, 2026):
    2,325 removal requests / ~141k verified = ~1.65% statewide.

    As more post-deadline snapshots accumulate, the observed rate gains
    credibility and the prior fades. Full credibility after clerk_window_days.
    """
    REMOVAL_PRIOR = 0.0165  # empirical prior: ~1.65% statewide removal rate
    credibility = min(1.0, days_post_deadline / clerk_window_days)
    return credibility * observed_rate + (1.0 - credibility) * REMOVAL_PRIOR


def compute_district_prob_survival(
    verified: float,
    threshold: int,
    peak_verified: int,           # highest count ever seen for this district
    post_deadline_removal_rate: float,  # removals / peak since submission deadline
    observed_removal_rate: float,       # full-history removal rate (background)
    days_remaining: int,
    days_post_deadline: int = 0,  # days elapsed since submission deadline
    pre_deadline_slope: float = 0.0,  # sigs/day velocity just before deadline
) -> float:
    """
    Post-submission-deadline survival model.

    No new signatures can be added. The question is purely:
    will enough of the current verified signatures survive clerk review?

    Uses a Bayesian blend of observed post-deadline removal rate with an
    empirical prior (1.65%) calibrated from county clerk removal-request
    data (KSL/SLTrib, Feb 12, 2026). Prior fades as more data accumulates.

    For below-threshold districts: also applies a trajectory multiplier —
    districts that were accelerating strongly just before the deadline are
    more likely to have unposted signatures still in the LG pipeline.

    Returns a value in [0.0, 1.0].
    """
    # Effective removal rate: Bayesian blend of observed + prior
    effective_rate = bayesian_removal_rate(post_deadline_removal_rate, days_post_deadline)

    if verified >= threshold:
        # Already met — P = 1.0 only if we think removals won't push below threshold
        buffer = (verified - threshold) / threshold
        if buffer >= 0.10:
            # Well above threshold — apply removal risk to see if we could fall back
            projected_remaining = verified * (1.0 - effective_rate)
            if projected_remaining >= threshold:
                return 1.0
            # Blended removal risk could theoretically breach threshold
            removal_risk = min(effective_rate * 2, 0.10)
            return max(0.92, 1.0 - removal_risk)
        # Close to threshold — meaningful chance of falling back below due to removals
        removal_risk = min(effective_rate * 3, 0.15)
        return max(0.90, 1.0 - removal_risk)

    # Below threshold — the gap tells us how likely late LG postings can close it.
    current_pct = verified / threshold if threshold > 0 else 0.0
    gap_pct = 1.0 - current_pct

    # Base gap-to-probability lookup (calibrated to 1.65% prior removal environment)
    if gap_pct <= 0.02:
        base_p = 0.18   # Within 2% — small chance LG posting lag resolves it
    elif gap_pct <= 0.05:
        base_p = 0.10
    elif gap_pct <= 0.10:
        base_p = 0.05
    elif gap_pct <= 0.15:
        base_p = 0.02
    elif gap_pct <= 0.25:
        base_p = 0.005
    else:
        return 0.00  # Structurally impossible

    # Trajectory multiplier: if a district had strong pre-deadline velocity,
    # it may have unposted signatures still in the LG pipeline. Boost base_p
    # proportional to the slope, capped at 1.5×. Only applies early post-deadline
    # while the LG lag window is still plausible (within 10 days of deadline).
    if pre_deadline_slope > 0 and days_post_deadline <= 10:
        # Normalize slope: 50 sigs/day = meaningful surge for any district
        trajectory_bonus = min(0.5, pre_deadline_slope / 50.0) * (1.0 - days_post_deadline / 10.0)
        base_p = base_p * (1.0 + trajectory_bonus)

    # Downward nudge if effective removal rate is meaningfully above prior
    # (i.e., this district is under heavier-than-average clerk scrutiny)
    if effective_rate > 0.03:
        base_p *= max(0.5, 1.0 - (effective_rate - 0.03) * 10)

    return round(max(0.0, min(0.99, base_p)), 4)


def compute_distribution(probs: list[float]) -> list[float]:
    """Exact DP: returns dp[k] = P(exactly k districts meet threshold)."""
    n = len(probs)
    dp = [0.0] * (n + 1)
    dp[0] = 1.0
    for p in probs:
        new_dp = [0.0] * (n + 1)
        for k in range(n + 1):
            if dp[k] == 0:
                continue
            new_dp[k + 1] += dp[k] * p
            new_dp[k] += dp[k] * (1 - p)
        dp = new_dp
    return dp


def p_qualify(dp: list[float]) -> float:
    return sum(dp[26:])


def expected_districts(probs: list[float]) -> float:
    return sum(probs)


# ---------------------------------------------------------------------------
# Trend calculation (used when history is absent)
# ---------------------------------------------------------------------------

def compute_trend(weekly: list[int]) -> str:
    """ACCEL / STABLE / DECEL based on last 2 vs prior 2 weeks."""
    if len(weekly) < 4:
        return "STABLE"
    last2 = sum(weekly[-2:])
    prior2 = sum(weekly[-4:-2])
    if prior2 == 0:
        return "STABLE"
    ratio = last2 / prior2
    if ratio >= 1.15:
        return "ACCEL"
    if ratio <= 0.85:
        return "DECEL"
    return "STABLE"


def compute_trend_from_history(district_snapshots: list[dict]) -> str:
    """
    Compute ACCEL/STABLE/DECEL from inter-snapshot deltas.

    Uses a linear regression on per-day rates (velocity series) to detect
    whether velocity is meaningfully increasing or decreasing. Requires the
    slope to exceed a noise threshold before calling ACCEL or DECEL, to
    avoid false signals from single-interval noise with small sample sizes.
    """
    if len(district_snapshots) < 3:
        return "STABLE"

    # Compute per-day rates for each interval
    rates = []
    xs = []  # interval index (for regression)
    for i in range(1, len(district_snapshots)):
        prev_date = date.fromisoformat(district_snapshots[i - 1]["date"])
        cur_date = date.fromisoformat(district_snapshots[i]["date"])
        days = max((cur_date - prev_date).days, 1)
        net = district_snapshots[i]["count"] - district_snapshots[i - 1]["count"]
        rates.append(net / days)  # sigs/day
        xs.append(i)

    if len(rates) < 2:
        return "STABLE"

    # Fit a linear regression to the velocity series
    # slope > 0 → velocity accelerating; slope < 0 → decelerating
    n = len(rates)
    x_mean = sum(xs) / n
    r_mean = sum(rates) / n
    num = sum((xs[i] - x_mean) * (rates[i] - r_mean) for i in range(n))
    den = sum((xs[i] - x_mean) ** 2 for i in range(n))
    slope = num / den if den != 0 else 0.0

    # Noise threshold: only signal ACCEL/DECEL if the slope exceeds 1 sig/day per interval
    # relative to the mean rate (i.e., >~10% change in rate per interval step).
    # This prevents noise from single-interval spikes dominating the signal.
    mean_abs_rate = max(abs(r_mean), 1.0)
    relative_slope = slope / mean_abs_rate  # normalized slope

    if relative_slope >= 0.10:
        return "ACCEL"
    if relative_slope <= -0.10:
        return "DECEL"
    return "STABLE"


# ---------------------------------------------------------------------------
# Tier classification
# ---------------------------------------------------------------------------

def classify_tier(prob: float, pct_verified: float = 0.0) -> str:
    if prob >= 1.0:
        return "CONFIRMED"
    if prob >= 0.90:
        return "NEARLY CERTAIN"
    if prob >= 0.70:
        return "VERY LIKELY"
    if prob >= 0.50:
        return "LIKELY"
    if prob >= 0.25:
        return "POSSIBLE"
    # Floors based on how far along a district already is:
    # ≥80% of threshold verified → at least LIKELY
    if pct_verified >= 0.80:
        return "LIKELY"
    # ≥60% of threshold verified → at least POSSIBLE
    if pct_verified >= 0.60:
        return "POSSIBLE"
    if prob >= 0.10:
        return "UNLIKELY"
    return "NO CHANCE"


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------

def parse_xlsx(path: Path) -> tuple[dict[int, int], dict[int, list[datetime]], dict[int, list[str]]]:
    """
    Read xlsx, return:
      - district_counts: {district_num: verified_count}
      - district_dates:  {district_num: [datetime, ...]}
      - district_names:  {district_num: [raw name string, ...]}
    """
    print(f"Reading {path} ...")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    district_counts: dict[int, int] = defaultdict(int)
    district_dates: dict[int, list[datetime]] = defaultdict(list)
    district_names: dict[int, list[str]] = defaultdict(list)
    skipped = 0
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 4:
            skipped += 1
            continue

        entry_date_raw = row[1]
        name_raw = row[2]
        district_raw = row[3]

        try:
            district = int(district_raw)
        except (TypeError, ValueError):
            skipped += 1
            continue

        if district not in THRESHOLDS:
            skipped += 1
            continue

        entry_dt = None
        if entry_date_raw:
            try:
                if isinstance(entry_date_raw, datetime):
                    entry_dt = entry_date_raw
                else:
                    entry_dt = dateutil_parser.parse(str(entry_date_raw))
            except Exception:
                entry_dt = None

        district_counts[district] += 1
        if entry_dt:
            district_dates[district].append(entry_dt)
        if name_raw:
            district_names[district].append(str(name_raw))
        total += 1

    wb.close()
    print(f"Parsed {total:,} rows ({skipped} skipped).")
    return dict(district_counts), dict(district_dates), dict(district_names)


# ---------------------------------------------------------------------------
# Weekly buckets (intra-file sparkline, used when no history)
# ---------------------------------------------------------------------------

def build_weekly_buckets(dates: list[datetime], n_buckets: int = WEEKLY_BUCKETS) -> list[int]:
    if not dates:
        return [0] * n_buckets
    sorted_dates = sorted(dates)
    earliest = sorted_dates[0]
    latest = sorted_dates[-1]
    span_days = max((latest - earliest).days, 1)
    bucket_size_days = span_days / n_buckets
    buckets = [0] * n_buckets
    for dt in sorted_dates:
        offset = (dt - earliest).days
        idx = min(int(offset / bucket_size_days), n_buckets - 1)
        buckets[idx] += 1
    return buckets


def build_weekly_buckets_from_history(district_snapshots: list[dict], n_buckets: int = WEEKLY_BUCKETS) -> list[int]:
    """
    Build WEEKLY_BUCKETS-length sparkline from inter-snapshot net deltas.
    Each entry represents net new signatures in that interval.
    """
    if not district_snapshots:
        return [0] * n_buckets

    # Compute deltas between consecutive snapshots
    deltas = []
    for i in range(1, len(district_snapshots)):
        net = max(0, district_snapshots[i]["count"] - district_snapshots[i - 1]["count"])
        deltas.append(net)

    if not deltas:
        return [0] * n_buckets

    # Pad or truncate to n_buckets
    if len(deltas) >= n_buckets:
        return deltas[-n_buckets:]
    else:
        return ([0] * (n_buckets - len(deltas))) + deltas


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Process petition xlsx → public/data.json")
    parser.add_argument("--file", help="Path to xlsx file (default: data/latest.xlsx)")
    args = parser.parse_args()

    xlsx_path = Path(args.file) if args.file else LATEST_PATH
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    source_file = xlsx_path.name

    # --- Load history ---
    history = load_history()
    if history:
        print(f"Loaded history.json: {history['snapshotCount']} snapshots, "
              f"{history['firstSnapshot']} → {history['lastSnapshot']}")
    else:
        print("No history.json found — using intra-file date bucketing only.")

    # --- Load previous data.json for deltas ---
    prev_data: dict = {}
    if DATA_JSON_PATH.exists():
        try:
            with open(DATA_JSON_PATH) as f:
                prev_data = json.load(f)
        except Exception as e:
            print(f"Warning: could not load previous data.json: {e}")

    prev_district_map: dict[int, dict] = {}
    if "districts" in prev_data:
        for d in prev_data["districts"]:
            prev_district_map[d["d"]] = d
    prev_overall = prev_data.get("overall", {})
    prev_p_qualify = prev_overall.get("pQualify", 0.0)
    prev_expected_districts = prev_overall.get("expectedDistricts", 0.0)

    # --- Parse xlsx ---
    district_counts, district_dates, district_names = parse_xlsx(xlsx_path)
    total_verified = sum(district_counts.values())

    # --- Detect reprocessing (same data as prev_data) ---
    # When CI re-runs on the same xlsx, verified counts match prev_data exactly.
    # In that case, carry forward the previous prevVerified/delta instead of zeroing.
    prev_total = prev_data.get("meta", {}).get("totalVerified", None)
    is_reprocessing = (prev_total is not None and prev_total == total_verified)
    if is_reprocessing:
        print(f"Reprocessing detected (total={total_verified} unchanged) — carrying forward previous deltas")

    # --- Determine model mode ---
    today = date.today()
    post_deadline = today > SUBMISSION_DEADLINE
    model_mode = "survival" if post_deadline else "growth"
    print(f"Model mode: {model_mode.upper()} "
          f"({'submission deadline passed' if post_deadline else 'accepting new signatures'})")

    # --- Build per-district history lookup (if available) ---
    district_history: dict[int, list[dict]] = {}
    rejection_rates: dict[int, float] = {}
    post_deadline_rates: dict[int, float] = {}
    peak_verified_map: dict[int, int] = {}
    projections: dict[int, dict] = {}
    daily_velocity = 0.0
    days_to_deadline = max((CLERK_DEADLINE - today).days, 0)

    if history:
        for d_num in THRESHOLDS:
            district_history[d_num] = [
                {"date": snap["date"], "count": snap["districts"].get(str(d_num), 0)}
                for snap in history["snapshots"]
            ]
        rejection_rates = {int(k): v for k, v in history.get("rejectionRates", {}).items()}
        post_deadline_rates = {int(k): v for k, v in history.get("postDeadlineRemovalRates", {}).items()}
        peak_verified_map = {int(k): v for k, v in history.get("peakVerified", {}).items()}
        projections = history.get("projections", {}).get("byDistrict", {})
        daily_velocity = history.get("dailyVelocity", 0.0)
        days_to_deadline = history.get("daysToDeadline", days_to_deadline)

    # --- Build per-district records ---
    districts_out = []
    all_probs = []
    all_growth_probs = []

    for d_num in sorted(THRESHOLDS.keys()):
        threshold = THRESHOLDS[d_num]
        verified = district_counts.get(d_num, 0)
        dates = district_dates.get(d_num, [])

        prev_rec = prev_district_map.get(d_num, {})
        if is_reprocessing:
            # Same data — carry forward previous deltas instead of zeroing
            prev_verified = prev_rec.get("prevVerified", verified)
            delta = prev_rec.get("delta", 0)
        else:
            prev_verified = prev_rec.get("verified", verified)
            delta = verified - prev_verified

        pct_verified = verified / threshold if threshold > 0 else 0.0

        # --- Trend & sparkline ---
        if history and d_num in district_history and len(district_history[d_num]) >= 2:
            hist_series = district_history[d_num]
            trend = compute_trend_from_history(hist_series)
            weekly = build_weekly_buckets_from_history(hist_series, WEEKLY_BUCKETS)
        else:
            weekly = build_weekly_buckets(dates, WEEKLY_BUCKETS)
            trend = compute_trend(weekly)

        final_week_sigs = weekly[-1] if weekly else 0

        # --- Rejection rates ---
        rejection_rate = rejection_rates.get(d_num, 0.0)
        post_deadline_rate = post_deadline_rates.get(d_num, 0.0)
        peak_verified = peak_verified_map.get(d_num, verified)

        # --- Probability (mode-dependent) ---
        if post_deadline:
            # SURVIVAL MODE: submission deadline passed.
            #
            # Key nuance: the LG office posts validations with a 1-2 business day lag.
            # Signatures submitted right before the Feb 15 deadline may not yet appear
            # in the LG count. The growth projection from history gives an upper bound
            # on what those late-posted signatures might add.
            #
            # We blend: early post-deadline → lean on growth projection as "pending"
            #           later → trust only the actual LG count
            #
            # LG_LAG_DAYS: how many calendar days we expect lag to persist (extended to 14)
            # Empirical: 22,934 net gains appeared Feb 16-20 (25.8% of pre-deadline total),
            # confirming the LG was still posting a large pre-deadline backlog through day 7.
            LG_LAG_DAYS = 14  # extended from 7 — data shows lag persists beyond one week
            days_elapsed = max(0, (date.today() - SUBMISSION_DEADLINE).days)
            # lag_weight decays from 1.0 on day 0 to 0.0 after LG_LAG_DAYS
            lag_weight = max(0.0, 1.0 - days_elapsed / LG_LAG_DAYS)

            # Growth projection upper bound from history
            growth_proj_raw = None
            if history and str(d_num) in projections:
                growth_proj_raw = projections[str(d_num)]["raw"]

            # Empirical LG lag estimate: observe net gains in post-deadline snapshots.
            # Any net additions after Feb 15 are pre-deadline sigs that LG hadn't posted yet.
            # If post-deadline gains are substantial (>0.5% of last pre-deadline count),
            # the lag window is still active; otherwise it has likely resolved.
            empirical_lag_active = True
            if history and "snapshots" in history:
                post_snaps = [
                    s for s in history["snapshots"]
                    if date.fromisoformat(s["date"]) > SUBMISSION_DEADLINE
                ]
                if len(post_snaps) >= 2:
                    # Total net additions (statewide) across post-deadline snapshots
                    post_gains = sum(
                        max(0, post_snaps[i]["total"] - post_snaps[i - 1]["total"])
                        for i in range(1, len(post_snaps))
                    )
                    last_pre_snap = next(
                        (s for s in reversed(history["snapshots"])
                         if date.fromisoformat(s["date"]) <= SUBMISSION_DEADLINE),
                        None
                    )
                    if last_pre_snap and last_pre_snap["total"] > 0:
                        gain_rate = post_gains / last_pre_snap["total"]
                        # If post-deadline gains are tiny (<0.1%), lag has likely resolved
                        empirical_lag_active = gain_rate >= 0.001

            # Empirical override: if post-deadline gains are still substantial (>5% of
            # pre-deadline total), the LG backlog is clearly not resolved — extend lag weight.
            # Conversely, if gains are tiny, lag has resolved and we trust LG counts fully.
            if not empirical_lag_active:
                lag_weight = 0.0
            elif lag_weight == 0.0 and empirical_lag_active:
                # Fixed window expired but data shows lag is still active — give a small floor
                lag_weight = 0.10

            # Per-district post-deadline velocity: sigs/day added or removed since Feb 15.
            # This is a real-time signal — if LG is still posting backlog sigs for this
            # district, the velocity will be positive. If clerk is removing, negative.
            post_deadline_velocity = 0.0  # sigs/day since deadline
            post_deadline_projected = float(verified)  # where this district lands by March 9
            if d_num in district_history:
                post_snaps = [
                    s for s in district_history[d_num]
                    if date.fromisoformat(s["date"]) > SUBMISSION_DEADLINE
                ]
                if len(post_snaps) >= 2:
                    # Compute velocity across all post-deadline intervals
                    total_net = post_snaps[-1]["count"] - post_snaps[0]["count"]
                    total_days = max(
                        1,
                        (date.fromisoformat(post_snaps[-1]["date"])
                         - date.fromisoformat(post_snaps[0]["date"])).days
                    )
                    post_deadline_velocity = total_net / total_days
                    # Project to March 9 using this velocity
                    post_deadline_projected = verified + post_deadline_velocity * days_to_deadline
                    post_deadline_projected = max(float(verified), post_deadline_projected) \
                        if post_deadline_velocity < 0 else post_deadline_projected
                elif len(post_snaps) == 1 and d_num in district_history:
                    # Only one post-deadline snapshot: use velocity from last pre→post interval
                    pre_snaps = [
                        s for s in district_history[d_num]
                        if date.fromisoformat(s["date"]) <= SUBMISSION_DEADLINE
                    ]
                    if pre_snaps:
                        span_days = max(1, (
                            date.fromisoformat(post_snaps[0]["date"])
                            - date.fromisoformat(pre_snaps[-1]["date"])
                        ).days)
                        post_deadline_velocity = (
                            post_snaps[0]["count"] - pre_snaps[-1]["count"]
                        ) / span_days
                        post_deadline_projected = verified + post_deadline_velocity * days_to_deadline

            # Effective verified count: use post-deadline projection when we have velocity data,
            # otherwise fall back to growth-projection blend weighted by lag.
            if len([s for s in district_history.get(d_num, [])
                    if date.fromisoformat(s["date"]) > SUBMISSION_DEADLINE]) >= 2:
                # Have real post-deadline velocity — use it directly, clipped to growth proj
                upper_bound = growth_proj_raw if growth_proj_raw else post_deadline_projected
                effective_verified = min(post_deadline_projected, upper_bound)
                effective_verified = max(float(verified), effective_verified)
            elif growth_proj_raw and lag_weight > 0:
                # No post-deadline velocity yet — blend with growth projection
                effective_verified = verified + lag_weight * max(0, growth_proj_raw - verified)
            else:
                effective_verified = float(verified)

            # Days elapsed since submission deadline (for Bayesian prior credibility)
            days_post_deadline = max(0, (date.today() - SUBMISSION_DEADLINE).days)

            # Pre-deadline slope: sigs/day rate in the last pre-deadline interval
            # Used to boost survival odds for districts that were surging just before cutoff
            pre_deadline_slope = 0.0
            if d_num in district_history:
                pre_snaps = [
                    s for s in district_history[d_num]
                    if date.fromisoformat(s["date"]) <= SUBMISSION_DEADLINE
                ]
                if len(pre_snaps) >= 2:
                    last_pre = pre_snaps[-1]
                    prev_pre = pre_snaps[-2]
                    interval_days = max(
                        1,
                        (date.fromisoformat(last_pre["date"]) - date.fromisoformat(prev_pre["date"])).days
                    )
                    pre_deadline_slope = max(0.0, (last_pre["count"] - prev_pre["count"]) / interval_days)

            # Pure survival prob from the current (possibly lag-blended) count
            survival_prob = compute_district_prob_survival(
                verified=effective_verified,
                threshold=threshold,
                peak_verified=peak_verified,
                post_deadline_removal_rate=post_deadline_rate,
                observed_removal_rate=rejection_rate,
                days_remaining=days_to_deadline,
                days_post_deadline=days_post_deadline,
                pre_deadline_slope=pre_deadline_slope,
            )

            # Growth prob: what would the growth model say about this district?
            # Used to blend during the lag window — if trajectory was strong, honor it.
            growth_prob_for_blend = compute_district_prob(
                verified=verified,
                threshold=threshold,
                trend=trend,
                final_week_sigs=final_week_sigs,
                projected_adj=growth_proj_raw if growth_proj_raw else float(verified),
                rejection_rate=rejection_rate,
            )

            # Blend: early post-deadline leans on growth signal;
            # as lag expires, survival signal takes over entirely.
            prob = lag_weight * growth_prob_for_blend + (1.0 - lag_weight) * survival_prob

            # Also compute a pure growth prob shadow (ignores survival mode, for toggle)
            # Projection: expected final count after remaining removals
            effective_rate = post_deadline_rate if post_deadline_rate > 0 else rejection_rate
            clerk_window_days = 20
            days_fraction = min(days_to_deadline / clerk_window_days, 1.0)
            projected_total = effective_verified - (peak_verified * effective_rate * days_fraction)
            projected_total = max(projected_total, float(verified))  # can't add, only remove
            projected_raw = growth_proj_raw if growth_proj_raw else projected_total
            projected_pct = projected_total / threshold if threshold > 0 else 0.0
        else:
            # GROWTH MODE: sigs still coming in
            if history and str(d_num) in projections:
                proj_data = projections[str(d_num)]
                projected_total = proj_data["rejectionAdjusted"]
                projected_pct = proj_data["pctOfThreshold"]
                projected_raw = proj_data["raw"]
            else:
                if total_verified > 0:
                    district_share = verified / total_verified
                else:
                    district_share = 1 / TOTAL_DISTRICTS
                surge_est = district_share * ESTIMATED_VALID_UNVERIFIED
                projected_total = verified + surge_est
                projected_raw = projected_total
                projected_pct = projected_total / threshold if threshold > 0 else 0.0

            prob = compute_district_prob(
                verified, threshold, trend, final_week_sigs,
                projected_total, rejection_rate
            )

        if is_reprocessing:
            prev_prob = prev_rec.get("prevProb", prob)
            prob_delta = prev_rec.get("probDelta", 0.0)
        else:
            prev_prob = prev_rec.get("prob", prob)
            prob_delta = round(prob - prev_prob, 4)
        tier = classify_tier(prob, pct_verified)
        all_probs.append(prob)

        # --- Growth-model shadow probability (always computed for toggle) ---
        if post_deadline:
            # What would the growth model say if we hadn't hit the deadline?
            growth_prob = compute_district_prob(
                verified, threshold, trend, final_week_sigs,
                projected_raw, rejection_rate
            )
        else:
            growth_prob = prob  # Already in growth mode

        all_growth_probs.append(growth_prob)

        districts_out.append({
            "d": d_num,
            "threshold": threshold,
            "verified": verified,
            "prevVerified": prev_verified,
            "delta": delta,
            "pctVerified": round(pct_verified, 4),
            "projectedTotal": round(projected_total, 1),
            "projectedRaw": round(projected_raw, 1),
            "projectedPct": round(projected_pct, 4),
            "peakVerified": peak_verified,
            "rejectionRate": round(rejection_rate, 4),
            "postDeadlineRate": round(post_deadline_rate, 4),
            "prob": round(prob, 4),
            "growthProb": round(growth_prob, 4),
            "prevProb": round(prev_prob, 4),
            "probDelta": round(prob_delta, 4),
            "tier": tier,
            "trend": trend,
            "weeklySignatures": weekly,
        })

    # --- DP distribution (primary model) ---
    dp = compute_distribution(all_probs)
    p_qual_raw = p_qualify(dp)
    exp_districts = expected_districts(all_probs)
    p_exact = [round(dp[k], 6) for k in range(TOTAL_DISTRICTS + 1)]

    # --- Correlation penalty: districts are not independent ---
    # Shared risks (statewide fraud wave, clerk ruling, organizing surge) create
    # inter-district correlation ~0.10-0.15. This causes the independence DP to
    # underestimate tail risk (p_qualify is too high). We apply a conservative
    # deflator that grows with p_qual_raw — the higher the raw probability, the
    # more a correlated downside matters. Penalty calibrated to ~2-3 pp at p=0.95.
    # Source: Utah county removal-request data (Feb 12) showed concentrated
    # fraud risk in specific districts (Utah Co., SL Co.), consistent with
    # correlated district-level exposure.
    CORRELATION_PENALTY_SCALE = 0.030  # max penalty at p_qual = 1.0
    correlation_penalty = CORRELATION_PENALTY_SCALE * p_qual_raw
    p_qual = max(0.0, p_qual_raw - correlation_penalty)

    # --- DP distribution (growth-model shadow, for toggle) ---
    dp_growth = compute_distribution(all_growth_probs)
    p_qual_growth_raw = p_qualify(dp_growth)
    correlation_penalty_growth = CORRELATION_PENALTY_SCALE * p_qual_growth_raw
    p_qual_growth = max(0.0, p_qual_growth_raw - correlation_penalty_growth)
    exp_districts_growth = expected_districts(all_growth_probs)
    p_exact_growth = [round(dp_growth[k], 6) for k in range(TOTAL_DISTRICTS + 1)]

    # --- Snapshot deltas ---
    gains = [d for d in districts_out if d["delta"] > 0]
    losses = [d for d in districts_out if d["delta"] < 0]
    biggest_gains = sorted(
        [{"d": d["d"], "delta": d["delta"], "verified": d["verified"], "threshold": d["threshold"]}
         for d in gains],
        key=lambda x: x["delta"], reverse=True
    )[:5]
    biggest_losses = sorted(
        [{"d": d["d"], "delta": d["delta"], "verified": d["verified"], "threshold": d["threshold"]}
         for d in losses],
        key=lambda x: x["delta"]
    )[:5]
    newly_met = [d["d"] for d in districts_out
                 if d["verified"] >= d["threshold"] and d["prevVerified"] < d["threshold"]]
    newly_failed = [d["d"] for d in districts_out
                    if d["verified"] < d["threshold"] and d["prevVerified"] >= d["threshold"]]
    if is_reprocessing:
        overall_prob_delta = prev_data.get("snapshot", {}).get("overallProbDelta", 0.0)
        expected_districts_delta = prev_data.get("snapshot", {}).get("expectedDistrictsDelta", 0.0)
    else:
        overall_prob_delta = round(p_qual - prev_p_qualify, 4)
        expected_districts_delta = round(exp_districts - prev_expected_districts, 2)

    # --- Anomalies from history —-- feed back into rejection rates ---
    # Districts with flagged packet-level fraud anomalies get a rejection rate bump.
    # A single anomaly adds +1 pp to their effective removal rate, capped at 5%.
    anomalies = history.get("anomalies", []) if history else []
    anomaly_districts = set(a["district"] for a in anomalies)
    for d_rec in districts_out:
        if d_rec["d"] in anomaly_districts:
            d_rec["rejectionRate"] = round(min(0.05, d_rec["rejectionRate"] + 0.01), 4)

    # --- Signature flow from history (net new / removals) ---
    if history and "snapshots" in history:
        snaps = history["snapshots"]
        last_snap = snaps[-1] if snaps else {}
        # Last interval
        interval_net = last_snap.get("totalDelta", 0)
        interval_removals = last_snap.get("totalRemovals", 0)
        # All-time totals
        alltime_added = sum(s.get("totalDelta", 0) for s in snaps if s.get("totalDelta", 0) > 0)
        alltime_removals = sum(s.get("totalRemovals", 0) for s in snaps)
        # Per-district removals this interval
        interval_district_removals = []
        if interval_removals > 0:
            for d_str, count in last_snap.get("removals", {}).items():
                if count > 0:
                    interval_district_removals.append({
                        "d": int(d_str),
                        "removed": count,
                    })
            interval_district_removals.sort(key=lambda x: x["removed"], reverse=True)
    else:
        interval_net = sum(d["delta"] for d in districts_out)
        interval_removals = 0
        alltime_added = interval_net
        alltime_removals = 0
        interval_district_removals = []

    # --- Statewide trajectory ---
    statewide_proj_raw = sum(d["projectedRaw"] for d in districts_out)
    statewide_proj_adj = sum(d["projectedTotal"] for d in districts_out)
    statewide_rejection_rate = history.get("statewideRejectionRate", 0.0) if history else 0.0

    # --- Statewide threshold projection ---
    statewide_target = QUALIFICATION_THRESHOLD_STATEWIDE
    statewide_remaining = max(0, statewide_target - total_verified)
    statewide_pct_complete = total_verified / statewide_target if statewide_target > 0 else 0.0

    # Compute weighted net daily velocity from recent snapshots
    net_daily_velocity = 0.0
    if history and "snapshots" in history and len(history["snapshots"]) >= 2:
        snaps = history["snapshots"]
        # Use last 5 intervals (6 snapshots), weighted: recent intervals count more
        intervals = []
        for i in range(max(1, len(snaps) - 5), len(snaps)):
            prev_date = date.fromisoformat(snaps[i - 1]["date"])
            cur_date = date.fromisoformat(snaps[i]["date"])
            days = max((cur_date - prev_date).days, 1)
            net = snaps[i]["total"] - snaps[i - 1]["total"]
            intervals.append((net / days, days))

        if intervals:
            # Exponential weighting: most recent interval gets highest weight
            weights = [2 ** i for i in range(len(intervals))]
            total_weight = sum(w for w in weights)
            net_daily_velocity = sum(r * w for (r, _), w in zip(intervals, weights)) / total_weight

    # Projected crossing date
    projected_crossing_date = None
    days_to_crossing = None
    if statewide_remaining > 0 and net_daily_velocity > 0:
        days_to_crossing = math.ceil(statewide_remaining / net_daily_velocity)
        crossing = today + __import__('datetime').timedelta(days=days_to_crossing)
        if crossing <= CLERK_DEADLINE:
            projected_crossing_date = crossing.isoformat()
        else:
            projected_crossing_date = None  # won't make it before deadline
    elif statewide_remaining <= 0:
        days_to_crossing = 0
        projected_crossing_date = today.isoformat()

    # P(reaching statewide target)
    # Use projectedStatewideRaw and projectedStatewideAdjusted as upper/lower bounds
    # for the expected final statewide count.
    # Blend: 60% adjusted (conservative), 40% raw (optimistic)
    projected_final = 0.6 * statewide_proj_adj + 0.4 * statewide_proj_raw
    if total_verified >= statewide_target:
        p_reach_target = 1.0
    elif projected_final >= statewide_target * 1.05:
        # Projected well above target — high confidence
        margin = (projected_final - statewide_target) / statewide_target
        p_reach_target = min(0.95, 0.90 + margin * 0.5)
    elif projected_final >= statewide_target:
        # Projected above but within 5% margin — moderate confidence
        margin = (projected_final - statewide_target) / statewide_target
        p_reach_target = 0.50 + (margin / 0.05) * 0.40  # 0.50 → 0.90
    else:
        # Projected below target — scale by shortfall
        shortfall = (statewide_target - projected_final) / statewide_target
        if shortfall <= 0.05:
            p_reach_target = 0.30 + (1.0 - shortfall / 0.05) * 0.20  # 0.30 → 0.50
        elif shortfall <= 0.15:
            p_reach_target = 0.10 + (1.0 - (shortfall - 0.05) / 0.10) * 0.20  # 0.10 → 0.30
        else:
            p_reach_target = max(0.01, 0.10 * (1.0 - shortfall))

    # Factor in velocity variance from recent snapshots for confidence band
    if history and "snapshots" in history and len(history["snapshots"]) >= 4:
        snaps = history["snapshots"]
        recent_rates = []
        for i in range(max(1, len(snaps) - 5), len(snaps)):
            prev_date = date.fromisoformat(snaps[i - 1]["date"])
            cur_date = date.fromisoformat(snaps[i]["date"])
            days = max((cur_date - prev_date).days, 1)
            net = snaps[i]["total"] - snaps[i - 1]["total"]
            recent_rates.append(net / days)
        if len(recent_rates) >= 2 and net_daily_velocity > 0:
            rate_std = (sum((r - net_daily_velocity) ** 2 for r in recent_rates) / len(recent_rates)) ** 0.5
            cv = rate_std / net_daily_velocity  # coefficient of variation
            # High variance = less certainty in our projection
            variance_penalty = min(cv * 0.15, 0.10)
            p_reach_target = max(0.01, p_reach_target - variance_penalty)

    on_track = projected_crossing_date is not None or total_verified >= statewide_target

    statewide_projection = {
        "target": statewide_target,
        "current": total_verified,
        "pctComplete": round(statewide_pct_complete, 4),
        "remaining": statewide_remaining,
        "netDailyVelocity": round(net_daily_velocity, 0),
        "projectedFinalCount": round(projected_final, 0),
        "projectedCrossingDate": projected_crossing_date,
        "daysToProjectedCrossing": days_to_crossing,
        "pReachTarget": round(min(1.0, max(0.0, p_reach_target)), 4),
        "onTrack": on_track,
    }

    # --- Build output ---
    now_utc = datetime.now(timezone.utc)

    # lastUpdated = the date of the actual data file, not the processing time.
    # Priority: history["lastSnapshot"] > xlsx filename (YYYY-MM-DD.xlsx) > today.
    data_date_str = None
    if history and history.get("lastSnapshot"):
        data_date_str = history["lastSnapshot"]   # e.g. "2026-02-20"
    else:
        # Try to parse date from filename (data/snapshots/2026-02-20.xlsx)
        stem = xlsx_path.stem   # "2026-02-20" or "latest" etc.
        try:
            date.fromisoformat(stem)   # validates format
            data_date_str = stem
        except ValueError:
            pass
    if not data_date_str:
        data_date_str = now_utc.strftime("%Y-%m-%d")

    # Build an ISO timestamp at noon UTC on data_date_str (neutral, avoids TZ artifacts)
    data_datetime_iso = f"{data_date_str}T12:00:00Z"

    output = {
        "meta": {
            "lastUpdated": data_date_str,
            "lastUpdatedISO": data_datetime_iso,
            "processedAt": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "sourceFile": source_file,
            "totalVerified": total_verified,
            "daysToDeadline": days_to_deadline,
            "dailyVelocity": round(daily_velocity, 1),
            "statewideRejectionRate": statewide_rejection_rate,
            "modelMode": model_mode,
            "postDeadline": post_deadline,
            "submissionDeadline": SUBMISSION_DEADLINE.isoformat(),
            "clerkDeadline": CLERK_DEADLINE_STR,
            "qualificationThreshold": QUALIFICATION_THRESHOLD_STATEWIDE,
            "districtsRequired": DISTRICTS_REQUIRED,
            "totalDistricts": TOTAL_DISTRICTS,
            "snapshotCount": history["snapshotCount"] if history else 1,
            "historyRange": f"{history['firstSnapshot']} → {history['lastSnapshot']}" if history else "n/a",
            "snapshotDates": [s["date"] for s in history["snapshots"]] if history else [],
        },
        "overall": {
            "pQualify": round(p_qual, 4),
            "expectedDistricts": round(exp_districts, 2),
            "pExact": p_exact,
            "projectedStatewideRaw": round(statewide_proj_raw, 0),
            "projectedStatewideAdjusted": round(statewide_proj_adj, 0),
            "pQualifyGrowth": round(p_qual_growth, 4),
            "expectedDistrictsGrowth": round(exp_districts_growth, 2),
            "pExactGrowth": p_exact_growth,
            "statewideProjection": statewide_projection,
        },
        "districts": districts_out,
        "snapshot": {
            "biggestGains": biggest_gains,
            "biggestLosses": biggest_losses,
            "newlyMet": newly_met,
            "newlyFailed": newly_failed,
            "overallProbDelta": overall_prob_delta,
            "expectedDistrictsDelta": expected_districts_delta,
            "signatureFlow": {
                "intervalNet": interval_net,
                "intervalRemovals": interval_removals,
                "intervalGross": interval_net + interval_removals,
                "alltimeAdded": alltime_added,
                "alltimeRemovals": alltime_removals,
                "districtRemovals": interval_district_removals[:5],
            },
            "anomalies": anomalies,
        },
    }

    # --- Write outputs ---
    PUBLIC_DIR.mkdir(parents=True, exist_ok=True)

    with open(DATA_JSON_PATH, "w") as f:
        json.dump(output, f, indent=2)
    print(f"Wrote {DATA_JSON_PATH}")

    lookup = build_lookup_index(district_names)
    with open(LOOKUP_INDEX_PATH, "w") as f:
        json.dump(lookup, f, separators=(',', ':'))
    print(f"Wrote {LOOKUP_INDEX_PATH} ({lookup['count']:,} name hashes)")

    confirmed = sum(1 for d in districts_out if d["verified"] >= d["threshold"])
    print(
        f"Summary: {now_utc.strftime('%Y-%m-%d')} | "
        f"Verified: {total_verified:,} | "
        f"Districts meeting threshold: {confirmed}/{TOTAL_DISTRICTS} | "
        f"P(qualify): {p_qual:.1%} | "
        f"Rejection rate: {statewide_rejection_rate:.1%}"
    )


if __name__ == "__main__":
    main()
