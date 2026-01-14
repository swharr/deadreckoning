#!/usr/bin/env python3
"""
scraper.py — Downloads the petition xlsx from the Utah Lt. Governor's website.

Usage:
    python scripts/scraper.py [--debug]

Exits with code 1 on any failure.
"""

import argparse
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

LG_PAGE_URL = (
    "https://vote.utah.gov/repeal-of-the-independent-redistricting-commission-"
    "and-standards-act-direct-initiative-list-of-signers/"
)

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
SNAPSHOTS_DIR = DATA_DIR / "snapshots"
MANUAL_DIR = DATA_DIR / "manual"
LATEST_PATH = DATA_DIR / "latest.xlsx"


def count_rows(path: Path) -> int:
    """Count data rows in xlsx (excluding header)."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        count = max(0, ws.max_row - 1)  # subtract header row
        wb.close()
        return count
    except Exception:
        return 0


def find_manual_file() -> Path | None:
    """Return the newest .xlsx in data/manual/ if one exists."""
    candidates = sorted(MANUAL_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def manual_is_newer(manual_path: Path) -> bool:
    """Check whether the manual file is newer than data/latest.xlsx."""
    if not LATEST_PATH.exists():
        return True
    return manual_path.stat().st_mtime > LATEST_PATH.stat().st_mtime


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def scrape_xlsx_url(debug: bool = False) -> str:
    """Fetch the LG petition page and extract the xlsx download URL."""
    try:
        resp = requests.get(LG_PAGE_URL, headers=HEADERS, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"ERROR: Could not fetch LG page: {e}", file=sys.stderr)
        sys.exit(1)

    html = resp.text

    if debug:
        print(html)
        return ""

    soup = BeautifulSoup(html, "lxml")

    # Strategy 1: anchor with text "Download Spreadsheet"
    for a in soup.find_all("a", href=True):
        text = (a.get_text() or "").strip()
        href = a["href"]
        if "Download Spreadsheet" in text and href.endswith(".xlsx"):
            return href

    # Strategy 2: any anchor whose href ends in .xlsx
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.endswith(".xlsx"):
            return href

    print("ERROR: No xlsx download link found on LG page.", file=sys.stderr)
    print("Run with --debug to inspect the raw HTML.", file=sys.stderr)
    sys.exit(1)


def download_file(url: str, dest: Path) -> None:
    """Download url to dest, streaming."""
    try:
        with requests.get(url, headers=HEADERS, timeout=120, stream=True) as resp:
            resp.raise_for_status()
            dest.parent.mkdir(parents=True, exist_ok=True)
            with open(dest, "wb") as f:
                for chunk in resp.iter_content(chunk_size=65536):
                    f.write(chunk)
    except requests.RequestException as e:
        print(f"ERROR: Download failed from {url}: {e}", file=sys.stderr)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Scrape petition xlsx from LG site.")
    parser.add_argument("--debug", action="store_true", help="Dump page HTML to stdout and exit.")
    args = parser.parse_args()

    SNAPSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    MANUAL_DIR.mkdir(parents=True, exist_ok=True)

    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    snapshot_path = SNAPSHOTS_DIR / f"{today_str}.xlsx"

    # --- Check for manual file first ---
    manual_path = find_manual_file()
    if manual_path and manual_is_newer(manual_path):
        print(f"Manual file found: {manual_path.name} — using instead of scraping.")
        dest_snapshot = SNAPSHOTS_DIR / manual_path.name
        shutil.move(str(manual_path), str(dest_snapshot))
        shutil.copy2(str(dest_snapshot), str(LATEST_PATH))
        rows = count_rows(LATEST_PATH)
        print(f"Downloaded: {dest_snapshot.name} ({rows} rows)")
        return

    # --- Scrape from LG site ---
    if args.debug:
        scrape_xlsx_url(debug=True)
        return

    xlsx_url = scrape_xlsx_url(debug=False)
    print(f"Found xlsx URL: {xlsx_url}")

    print(f"Downloading to {snapshot_path} ...")
    download_file(xlsx_url, snapshot_path)

    # Copy to latest.xlsx
    shutil.copy2(str(snapshot_path), str(LATEST_PATH))

    rows = count_rows(LATEST_PATH)
    print(f"Downloaded: {snapshot_path.name} ({rows} rows)")


if __name__ == "__main__":
    main()
